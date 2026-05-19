"""Excel输出写入器"""

import os
from typing import List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .base import OutputWriter
from ..models import BiddingRecord, LinkStatus


class ExcelWriter(OutputWriter):
    """Excel输出写入器 - 在原始文件基础上追加分析结果列"""

    # 状态对应的样式
    STATUS_STYLES = {
        LinkStatus.VALID: ("有效", "00B050"),       # 绿色
        LinkStatus.NEEDS_LOGIN: ("需登录查看", "FFC000"),  # 橙色
        LinkStatus.INVALID: ("链接无效", "FF0000"),       # 红色
    }

    def write(self, records: List[BiddingRecord], output_path: str) -> str:
        """写入Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "招中标分析结果"

        # 写入表头
        headers = [
            "序号", "标讯标题", "原文链接", "来源网站名称",
            "招中标类型", "招中标阶段", "采购单位", "中标单位",
            "中标金额", "省份", "城市", "项目名称", "项目大类",
            "招标编号", "发布时间",
            "链接状态", "AI摘要", "页面摘要", "检测时间", "错误信息"
        ]

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, record in enumerate(records, start=2):
            row_data = [
                record.index,
                record.标讯标题,
                record.原文链接,
                record.来源网站名称,
                record.招中标类型,
                record.招中标阶段,
                record.采购单位,
                record.中标单位,
                record.中标金额 if record.中标金额 else "",
                record.省份,
                record.城市,
                record.项目名称,
                record.项目大类,
                record.招标编号,
                record.发布时间,
                record.链接状态.value,
                record.AI摘要,
                record.页面摘要,
                record.检测时间,
                record.错误信息,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 链接状态列着色
            status_cell = ws.cell(row=row_idx, column=16)
            if record.链接状态 in self.STATUS_STYLES:
                _, color = self.STATUS_STYLES[record.链接状态]
                status_cell.font = Font(color=color, bold=True)

            # AI摘要列着色（蓝色标记）
            ai_cell = ws.cell(row=row_idx, column=17)
            if record.AI摘要:
                ai_cell.font = Font(color="0070C0")

            # 原文链接添加超链接
            link_cell = ws.cell(row=row_idx, column=3)
            if record.原文链接:
                link_cell.hyperlink = record.原文链接
                link_cell.font = Font(color="0563C1", underline="single")

        # 设置列宽
        col_widths = {
            1: 6,   # 序号
            2: 35,  # 标讯标题
            3: 40,  # 原文链接
            4: 18,  # 来源网站名称
            5: 10,  # 招中标类型
            6: 12,  # 招中标阶段
            7: 25,  # 采购单位
            8: 25,  # 中标单位
            9: 12,  # 中标金额
            10: 8,  # 省份
            11: 10, # 城市
            12: 25, # 项目名称
            13: 12, # 项目大类
            14: 20, # 招标编号
            15: 12, # 发布时间
            16: 12, # 链接状态
            17: 50, # AI摘要
            18: 50, # 页面摘要（结构化信息）
            19: 18, # 检测时间
            20: 20, # 错误信息
        }

        for col, width in col_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        return os.path.abspath(output_path)
