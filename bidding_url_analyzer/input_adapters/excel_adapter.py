"""Excel输入适配器"""

import os
from typing import List
from openpyxl import load_workbook
from .base import InputAdapter
from ..models import BiddingRecord


# Excel列名到模型字段的映射
COLUMN_MAPPING = {
    "标讯标题": "标讯标题",
    "原文链接": "原文链接",
    "来源网站名称": "来源网站名称",
    "招中标类型": "招中标类型",
    "招中标阶段": "招中标阶段",
    "采购单位": "采购单位",
    "中标单位": "中标单位",
    "中标金额": "中标金额",
    "招标金额": "招标金额",
    "省份": "省份",
    "城市": "城市",
    "项目名称": "项目名称",
    "项目大类": "项目大类",
    "招标编号": "招标编号",
    "发布时间": "发布时间",
}


class ExcelAdapter(InputAdapter):
    """Excel文件输入适配器"""

    def load(self, source: str) -> List[BiddingRecord]:
        if not os.path.exists(source):
            raise FileNotFoundError(f"文件不存在: {source}")

        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        # 构建列索引映射
        col_map = {}
        for i, header in enumerate(headers):
            if header in COLUMN_MAPPING:
                col_map[header] = i

        records = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not row or not any(row):
                continue

            raw_data = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    raw_data[headers[i]] = val

            record = BiddingRecord(
                index=row_idx,
                raw_data=raw_data,
            )

            # 填充映射字段
            for excel_col, model_field in COLUMN_MAPPING.items():
                if excel_col in col_map and col_map[excel_col] < len(row):
                    val = row[col_map[excel_col]]
                    if model_field in ("中标金额", "招标金额"):
                        try:
                            val = float(val) if val else 0.0
                        except (ValueError, TypeError):
                            val = 0.0
                    else:
                        val = str(val) if val else ""
                    setattr(record, model_field, val)

            # 如果原文链接为空则跳过
            if not record.原文链接:
                continue

            records.append(record)

        wb.close()
        return records

    def get_record_count(self, source: str) -> int:
        if not os.path.exists(source):
            return 0
        wb = load_workbook(source, read_only=True, data_only=True)
        ws = wb.active
        count = ws.max_row - 1  # 减去表头
        wb.close()
        return max(0, count)
